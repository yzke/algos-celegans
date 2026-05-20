"""Load the Cook et al. 2019 hermaphrodite connectome into matrix form.

Phase 0 deliverable: produce a fully-typed `ConnectomeData` object with
- `W_chem` (302×302, signed and normalized; row = post-synaptic, col = pre-)
- `W_gap`  (302×302, symmetric, non-negative, normalized)
- per-neuron metadata (name, category, neurotransmitter)

The 302-neuron list is constructed from the Cook 2019 SI 5 spreadsheet itself:
  - All 300 row labels in the "hermaphrodite chemical" sheet (every neuron
    that emits at least one chemical synapse), plus
  - CANL and CANR — the two canonical hermaphrodite neurons that have no
    chemical out-synapses but appear in the gap-junction sheet.

This avoids hard-coding a separate "302 neurons" list and keeps the loader
re-runnable against any future correction of the Cook 2019 data.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import openpyxl

from algos.config import (
    COOK2019_XLSX,
    CONNECTOME_CACHE,
    N_NEURONS,
    NORMALIZATION_MODE,
)
from algos.neurotransmitters import neurotransmitter_sign

CHEM_SHEET = "hermaphrodite chemical"
GAP_SHEET = "hermaphrodite gap jn symmetric"

# Section markers found in column A of each sheet. Anything outside these
# section ranges (e.g. "MUSCLES", "SEX-SPECIFIC CELLS" with muscle cells) is
# excluded from the 302-neuron set.
NEURON_SECTIONS = {"PHARYNX", "SENSORY NEURONS", "INTERNEURONS", "MOTOR NEURONS"}
SEX_SPECIFIC_SECTION = "SEX SPECIFIC"
SEX_SPECIFIC_SHEET_SECTION = "SEX-SPECIFIC CELLS"

# Cells we explicitly want to keep from the sex-specific block in the gap
# sheet (the rest are vulval/uterine muscles which are not part of the 302).
SEX_SPECIFIC_NEURONS = {
    "HSNL", "HSNR",
    "VC01", "VC02", "VC03", "VC04", "VC05", "VC06",
}

# Section categories used as metadata. Keys are the sheet section labels;
# values are short canonical category names used elsewhere in the codebase.
SECTION_TO_CATEGORY = {
    "PHARYNX": "pharyngeal",
    "SENSORY NEURONS": "sensory",
    "INTERNEURONS": "interneuron",
    "MOTOR NEURONS": "motor",
    "SEX SPECIFIC": "sex_specific",
    "SEX-SPECIFIC CELLS": "sex_specific",
}


# ---------------------------------------------------------------------------
# Data class
# ---------------------------------------------------------------------------


@dataclass
class ConnectomeData:
    """Loaded connectome: matrices + per-neuron metadata.

    Notes on conventions:
      - `W_chem[i, j]` is the synapse from neuron `j` (pre) to neuron `i`
        (post). This matches design.md §3.1 and is the form used by
        `dynamics.neural_step` (which computes `W_chem @ sigmoid(V)`).
      - `W_gap` is symmetric and non-negative.
      - All weights are normalized so that `max(|W_chem|) <= 1` and
        `max(W_gap) <= 1` (see `algos.config.NORMALIZE_BY_GLOBAL_MAX`).
    """

    n_neurons: int
    W_chem: np.ndarray            # (N, N) float
    W_gap: np.ndarray             # (N, N) float, symmetric
    neuron_names: list[str]
    neuron_to_idx: dict[str, int]
    neurotransmitter: list[str]   # per-neuron, either "GABA" or "default"
    category: list[str]           # per-neuron category
    # Raw (unsigned, unnormalized) chemical counts — preserved for diagnostics.
    W_chem_raw: np.ndarray | None = None
    W_gap_raw: np.ndarray | None = None

    # ------------------------------------------------------------------ load

    @classmethod
    def load(
        cls,
        xlsx_path: Path | None = None,
        *,
        cache_path: Path | None = None,
        use_cache: bool = True,
    ) -> "ConnectomeData":
        """Load the connectome from the Cook 2019 xlsx, with a .npz cache.

        Args:
            xlsx_path: Path to ``SI 5 ... corrected July 2020.xlsx``. Defaults
                to ``algos.config.COOK2019_XLSX``.
            cache_path: Where to read/write the compiled `.npz` cache.
            use_cache: If True (default), read from the cache when present and
                write it on first compile.
        """
        xlsx_path = Path(xlsx_path) if xlsx_path is not None else COOK2019_XLSX
        cache_path = Path(cache_path) if cache_path is not None else CONNECTOME_CACHE

        if use_cache and cache_path.exists():
            return cls._from_cache(cache_path)

        if not xlsx_path.exists():
            raise FileNotFoundError(
                f"Cook 2019 connectome xlsx not found at {xlsx_path}. "
                "See data/connectome/README.md for download instructions."
            )

        data = cls._from_xlsx(xlsx_path)

        if use_cache:
            data.save_cache(cache_path)
        return data

    # ------------------------------------------------------------- queries

    def get_neuron_indices_by_category(self, category: str) -> list[int]:
        """Return indices of all neurons in the given category."""
        return [i for i, c in enumerate(self.category) if c == category]

    def idx(self, name: str) -> int:
        """Convenience: name → index, raising a clearer error on miss."""
        try:
            return self.neuron_to_idx[name]
        except KeyError as exc:
            raise KeyError(f"Unknown neuron name: {name!r}") from exc

    # ------------------------------------------------------------- caching

    def save_cache(self, path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        np.savez_compressed(
            path,
            W_chem=self.W_chem,
            W_gap=self.W_gap,
            W_chem_raw=self.W_chem_raw if self.W_chem_raw is not None
                else np.zeros((0, 0)),
            W_gap_raw=self.W_gap_raw if self.W_gap_raw is not None
                else np.zeros((0, 0)),
            neuron_names=np.array(self.neuron_names),
            neurotransmitter=np.array(self.neurotransmitter),
            category=np.array(self.category),
        )

    @classmethod
    def _from_cache(cls, path: Path) -> "ConnectomeData":
        npz = np.load(path, allow_pickle=False)
        names = [str(s) for s in npz["neuron_names"]]
        raw_chem = npz["W_chem_raw"]
        raw_gap = npz["W_gap_raw"]
        return cls(
            n_neurons=len(names),
            W_chem=npz["W_chem"],
            W_gap=npz["W_gap"],
            neuron_names=names,
            neuron_to_idx={n: i for i, n in enumerate(names)},
            neurotransmitter=[str(s) for s in npz["neurotransmitter"]],
            category=[str(s) for s in npz["category"]],
            W_chem_raw=raw_chem if raw_chem.size else None,
            W_gap_raw=raw_gap if raw_gap.size else None,
        )

    # ----------------------------------------------------------- xlsx core

    @classmethod
    def _from_xlsx(cls, xlsx_path: Path) -> "ConnectomeData":
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        chem_grid, chem_row_labels, chem_col_labels, chem_row_sections, chem_col_sections = (
            _read_matrix_sheet(wb[CHEM_SHEET])
        )
        gap_grid, gap_row_labels, gap_col_labels, gap_row_sections, _ = (
            _read_matrix_sheet(wb[GAP_SHEET])
        )

        # ---- Build canonical 302-neuron list ------------------------------
        # Start with every chemical row label (300 neurons that emit chem synapses).
        neuron_names: list[str] = list(chem_row_labels)

        # Add CANL/CANR (canonical members of the 302; appear in gap sheet
        # and as chemical *receivers* but never as senders).
        for can in ("CANL", "CANR"):
            if can in gap_row_labels and can not in neuron_names:
                neuron_names.append(can)

        if len(neuron_names) != N_NEURONS:
            raise ValueError(
                f"Expected {N_NEURONS} neurons, derived {len(neuron_names)}. "
                f"First few: {neuron_names[:5]}"
            )

        idx = {name: i for i, name in enumerate(neuron_names)}

        # ---- Per-neuron metadata ------------------------------------------
        category: list[str] = []
        for name in neuron_names:
            if name in ("CANL", "CANR"):
                # Cook 2019 files CANL/CANR under the gap sheet's MUSCLES
                # section because they have no chemical or canonical-position
                # placement. They are genuine neurons (excretory canal-
                # associated) so we tag them explicitly.
                category.append("other_neuron")
                continue
            sec = chem_row_sections.get(name) or gap_row_sections.get(name)
            category.append(SECTION_TO_CATEGORY.get(sec or "", "other"))

        neurotransmitter = [
            ("GABA" if neurotransmitter_sign(n) == -1 else "default")
            for n in neuron_names
        ]

        # ---- Build W_chem raw ---------------------------------------------
        # Cook 2019 convention: rows = pre, cols = post.
        # Our convention: W[i, j] = synapse from j (pre) to i (post).
        # → W[i, j] = sheet[row=j_label, col=i_label] (i.e. transpose of the sheet block).
        W_chem_raw = np.zeros((N_NEURONS, N_NEURONS), dtype=np.float64)
        for r_name, r_vals in chem_grid.items():
            j = idx.get(r_name)
            if j is None:
                continue  # pre-neuron not in our 302 list (shouldn't happen for chem)
            for c_name, val in r_vals.items():
                i = idx.get(c_name)
                if i is None:
                    continue  # post-cell is a muscle/glia/etc, skip
                W_chem_raw[i, j] = val

        # ---- Build W_gap raw ----------------------------------------------
        W_gap_raw = np.zeros((N_NEURONS, N_NEURONS), dtype=np.float64)
        for r_name, r_vals in gap_grid.items():
            i = idx.get(r_name)
            if i is None:
                continue
            for c_name, val in r_vals.items():
                j = idx.get(c_name)
                if j is None:
                    continue
                W_gap_raw[i, j] = val

        # Hard-enforce gap symmetry. The "symmetric" sheet should already be,
        # but rounding/quirks can leave 1-section asymmetries; take the elementwise max.
        W_gap_raw = np.maximum(W_gap_raw, W_gap_raw.T)
        # The Cook 2019 corrected gap sheet has a handful (~14) of non-zero
        # diagonal entries (self-gap-junctions). These are data artifacts —
        # a real gap junction is between two distinct cells, and our gap_input
        # formula `W_gap @ V - V * sum(W_gap)` algebraically cancels self
        # entries anyway. Zeroing the diagonal keeps `W_gap` topologically
        # clean and makes counts in tests/diagnostics meaningful.
        np.fill_diagonal(W_gap_raw, 0.0)

        # ---- Apply neurotransmitter signs to W_chem -----------------------
        signs = np.array(
            [neurotransmitter_sign(n) for n in neuron_names], dtype=np.float64
        )
        # Pre-synaptic neuron is column j; flipping all of j's outputs means
        # multiplying column j by signs[j].
        W_chem_signed = W_chem_raw * signs[np.newaxis, :]

        # ---- Normalize so the chem/gap inputs stay O(1) -------------------
        W_chem = _normalize(W_chem_signed, NORMALIZATION_MODE)
        W_gap = _normalize(W_gap_raw, NORMALIZATION_MODE)
        if NORMALIZATION_MODE == "per_row":
            # Re-enforce gap symmetry after independent row scaling.
            W_gap = 0.5 * (W_gap + W_gap.T)

        return cls(
            n_neurons=N_NEURONS,
            W_chem=W_chem,
            W_gap=W_gap,
            neuron_names=neuron_names,
            neuron_to_idx=idx,
            neurotransmitter=neurotransmitter,
            category=category,
            W_chem_raw=W_chem_raw,
            W_gap_raw=W_gap_raw,
        )


# ---------------------------------------------------------------------------
# Sheet reader
# ---------------------------------------------------------------------------


def _read_matrix_sheet(ws) -> tuple[
    dict[str, dict[str, float]],
    list[str],
    list[str],
    dict[str, str],
    dict[str, str],
]:
    """Parse a Cook 2019 SI 5 matrix sheet.

    Returns:
        grid: row_label → {col_label: value} for non-zero cells
        row_labels: row neuron names in sheet order
        col_labels: col neuron names in sheet order
        row_sections: row_label → section label (PHARYNX / SENSORY / ...)
        col_sections: col_label → section label
    """
    rows = list(ws.iter_rows(values_only=True))

    # Column labels live on sheet row index 2 (0-based), starting at col index 3.
    header_row = rows[2]
    col_labels = [str(c) for c in header_row[3:] if c is not None]

    # Section labels for columns: scan row 0; each section header marks the
    # *starting* sheet-column index for that section.
    section_row = rows[0]
    col_sections: dict[str, str] = {}
    current_section: str | None = None
    for sheet_col_idx, value in enumerate(section_row[3:], start=3):
        if value is not None:
            current_section = str(value).strip()
        label_in_header = header_row[sheet_col_idx]
        if label_in_header is not None and current_section is not None:
            col_sections[str(label_in_header)] = current_section

    # Row labels live in column 2 starting at sheet row index 3.
    row_labels: list[str] = []
    row_sections: dict[str, str] = {}
    current_section = None
    for sheet_row_idx in range(3, len(rows)):
        row = rows[sheet_row_idx]
        sec_label = row[0]
        if sec_label is not None:
            current_section = str(sec_label).strip()
        name = row[2]
        if name is None:
            continue
        name = str(name).strip()
        row_labels.append(name)
        if current_section is not None:
            row_sections[name] = current_section

    # Build the sparse grid (only non-empty numeric cells).
    grid: dict[str, dict[str, float]] = {}
    # Map sheet column index → label, ignoring None header cells.
    col_idx_to_label: dict[int, str] = {}
    for sheet_col_idx in range(3, ws.max_column):
        label = header_row[sheet_col_idx]
        if label is not None:
            col_idx_to_label[sheet_col_idx] = str(label).strip()

    for sheet_row_idx in range(3, len(rows)):
        row = rows[sheet_row_idx]
        name = row[2]
        if name is None:
            continue
        name = str(name).strip()
        cells: dict[str, float] = {}
        for sheet_col_idx, col_label in col_idx_to_label.items():
            if sheet_col_idx >= len(row):
                continue
            v = row[sheet_col_idx]
            if v is None or v == "":
                continue
            try:
                cells[col_label] = float(v)
            except (TypeError, ValueError):
                # Skip non-numeric cells; the spreadsheet has occasional notes.
                continue
        if cells:
            grid[name] = cells

    return grid, row_labels, col_labels, row_sections, col_sections


def _normalize(W: np.ndarray, mode: str) -> np.ndarray:
    """Scale a weight matrix to make per-row total input O(1)."""
    if mode == "none":
        return W.copy()
    if mode == "global_max":
        s = max(1.0, float(np.max(np.abs(W))))
        return W / s
    if mode == "per_row":
        row_l1 = np.abs(W).sum(axis=1, keepdims=True)
        row_scale = np.maximum(1.0, row_l1)
        return W / row_scale
    raise ValueError(f"Unknown normalization mode: {mode!r}")


__all__ = ["ConnectomeData"]
